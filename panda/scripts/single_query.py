import io
from django.http import FileResponse
from openpyxl import Workbook
from panda.models import Employee, Route, Driver
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned


def type_converter(value: str) -> str:
    classes = {
        "name":"nm",
        "driver_name":"dr",
        "address":"ad",
        "phone":"ph",
        "company":"cp",
        "comments":"cm",
        "shift":"sf",
    }
    return classes[value]

def get_single_route(data):
    lookup_value = data["lookup_value"]
    lookup_type = type_converter(data["lookup_type"])
    

    try:
        if lookup_type == "dr":
            driver_id = Driver.objects.get(id=lookup_value)
            routes = Route.objects.filter(driver_id=driver_id)
        elif lookup_type == "ad":
            routes = Route.objects.filter(address__icontains=lookup_value)
        elif lookup_type == "sf":
            routes = Route.objects.filter(shift__icontains=lookup_value)
        elif lookup_type == "dc":
            routes = Route.objects.filter(descr__icontains=lookup_value)
        elif lookup_type == "cp":
            routes = Route.objects.filter(company__icontains=lookup_value)
        else:
            # Handle unsupported lookup types
            return {
                "message": "Route not found | Lookup not supported",
                "error": True,
                "data": None,
            }

        # Create an in-memory Excel workbook
        wb = Workbook()
        ws = wb.active

        # Add headers to the Excel file
        headers = ["Nombre Driver", "Direccion", "Descripcion", "Turno", "Compania"]
        ws.append(headers)

        # Add data rows to the Excel file
        for route in routes:
            # Get the driver's name using the Driver model
            driver_name = route.driver_id.name if route.driver_id else ""
            route_data = route.to_object()
            row = [
                driver_name,
                route_data["address"],
                route_data["descr"],
                route_data["shift"],
                route_data["company"],
            ]
            ws.append(row)

        # Create a BytesIO object to save the workbook contents
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return the Excel file as a response
        response = FileResponse(excel_file, as_attachment=True, filename="routes.xlsx")
        return response
    except Exception as e:
        return {
            "message": f"Route not found | {e} | {e.__traceback__.tb_lineno}",
            "error": True,
            "data": None,
        }


def get_single_driver(data):
    lookup_value = data["lookup_value"]
    lookup_type = type_converter(data["lookup_type"])

    try:
        if lookup_type == "nm":
            drivers = Driver.objects.filter(name__icontains=lookup_value)
        elif lookup_type == "ad":
            drivers = Driver.objects.filter(address__icontains=lookup_value)
        elif lookup_type == "ph":
            drivers = Driver.objects.filter(phone__icontains=lookup_value)
        else:
            # Handle unsupported lookup types
            return {
                "message": "Driver not found | Lookup not supported",
                "error": True,
                "data": None,
            }

        # Create an in-memory Excel workbook
        wb = Workbook()
        ws = wb.active

        # Add headers to the Excel file
        headers = ["Nombre", "Direccion", "Telefono"]
        ws.append(headers)

        # Add data rows to the Excel file
        for driver in drivers:
            driver_data = driver.to_object()
            row = [driver_data["name"], driver_data["address"], driver_data["phone"]]
            ws.append(row)

        # Create a BytesIO object to save the workbook contents
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return the Excel file as a response
        response = FileResponse(excel_file, as_attachment=True, filename="drivers.xlsx")
        return response
    except Exception as e:
        return {
            "message": f"Driver not found | {e} | {e.__traceback__.tb_lineno}",
            "error": True,
            "data": None,
        }


def get_single_employee(data):
    lookup_value = data["lookup_value"]
    lookup_type = type_converter(data["lookup_type"])

    try:
        if lookup_type == "nm":
            employees = Employee.objects.filter(name__icontains=lookup_value)
        elif lookup_type == "ad":
            employees = Employee.objects.filter(address__icontains=lookup_value)
        elif lookup_type == "ph":
            employees = Employee.objects.filter(phone__icontains=lookup_value)
        elif lookup_type == "sf":
            employees = Employee.objects.filter(shift__icontains=lookup_value)
        elif lookup_type == "dr":
            driver_id = Driver.objects.get(name__icontains=lookup_value).id
            employees = Employee.objects.filter(driver_id=driver_id)
        elif lookup_type == "rt":
            route_id = Route.objects.get(descr=lookup_value)
            employees = Employee.objects.filter(route_id=route_id)
        elif lookup_type == "cp":
            employees = Employee.objects.filter(company__icontains=lookup_value)
        elif lookup_type == "cm":
            employees = Employee.objects.filter(comments__icontains=lookup_value)
        else:
            # Handle unsupported lookup types
            raise Exception("Employee not found | Lookup not supported")
        # Create an in-memory Excel workbook
        wb = Workbook()
        ws = wb.active

        # Add headers to the Excel file
        headers = [
            "Nombre",
            "Direccion",
            "Telefono",
            "Turno",
            "Driver",
            "Ruta",
            "Compania",
            "Comentarios",
        ]
        ws.append(headers)

        # Add data rows to the Excel file
        for employee in employees:
            driver_name = employee.driver_id.name if employee.driver_id else ""
            route_name = employee.route_id.descr if employee.route_id else ""
            employee_data = employee.to_object()
            row = [
                employee_data["name"],
                employee_data["address"],
                employee_data["phone"],
                employee_data["shift"],
                driver_name,
                route_name,
                employee_data["company"],
                employee_data["comments"],
            ]
            ws.append(row)

        # Create a BytesIO object to save the workbook contents
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return the Excel file as a response
        response = FileResponse(excel_file, as_attachment=True, filename="employee.xlsx")
        return response
    except Exception as e:
        return {
            "message": f"Driver not found | {e} | {e.__traceback__.tb_lineno}",
            "error": True,
            "data": None,
        }


single = {"route": get_single_route, "driver": get_single_driver, "employee": get_single_employee}
